# export.py — exports scored results to Excel with formatting

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from score import build_factor_table, score
from config import TICKERS, WEIGHTS

# ── Colors ────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "D9E1F2"
GREEN       = "E2EFDA"
YELLOW      = "FFEB9C"
RED         = "FFC7CE"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F2F2F2"

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def score_color(val):
    if pd.isna(val):
        return None
    if val >= 65:
        return GREEN
    elif val >= 40:
        return YELLOW
    else:
        return RED

def format_sheet(ws, df, title):
    # Title row
    ws.insert_rows(1)
    ws.insert_rows(1)
    ws["A1"] = title
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=WHITE)
    ws["A1"].fill = make_fill(DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A1:{get_column_letter(ws.max_column)}1")

    ws["A2"] = "For informational purposes only. Not investment advice."
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws.merge_cells(f"A2:{get_column_letter(ws.max_column)}2")

    # Header row (row 3)
    header_row = 3
    for cell in ws[header_row]:
        cell.font = Font(name="Arial", bold=True, color=WHITE, size=11)
        cell.fill = make_fill(MID_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = make_border()
    ws.row_dimensions[header_row].height = 20

    # Data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row+1, max_row=ws.max_row), start=0):
        bg = WHITE if row_idx % 2 == 0 else LIGHT_GRAY
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.fill = make_fill(bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = make_border()
            cell.number_format = "0.00"

        # Color code composite score (column B)
        composite_cell = row[1]
        color = score_color(composite_cell.value)
        if color:
            composite_cell.fill = make_fill(color)
            composite_cell.font = Font(name="Arial", bold=True, size=10)

    # Column widths
    col_widths = {
        "A": 10,  # ticker
        "B": 14,  # composite
        "C": 12,  # ev_ebit
        "D": 12,  # price_fcf
        "E": 10,  # roic
        "F": 16,  # gm_stability
        "G": 18,  # net_debt_ebitda
        "H": 11,  # n_factors
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze panes below header
    ws.freeze_panes = f"A{header_row+1}"

def add_methodology_tab(wb):
    ws = wb.create_sheet("Methodology")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 40

    # Title
    ws["A1"] = "Systematic Equity Factor Screener — Methodology"
    ws["A1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["A1"].fill = make_fill(DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:D1")

    # Overview
    ws["A3"] = "Overview"
    ws["A3"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
    ws["A4"] = "This model scores companies on 5 fundamental factors and ranks them by composite attractiveness."
    ws["A4"].font = Font(name="Arial", size=10)
    ws.merge_cells("A4:D4")

    # Factor table header
    headers = ["Factor", "Description", "Weight", "Direction"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = Font(name="Arial", bold=True, color=WHITE)
        cell.fill = make_fill(MID_BLUE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = make_border()

    # Factor rows — weights pulled from config.WEIGHTS so this sheet can't drift from the code
    factors = [
        ["EV/EBIT",          "Enterprise value relative to operating profit. Measures cheapness net of debt.", f"{WEIGHTS['ev_ebit']:.0%}", "Lower = Better"],
        ["Price/FCF",        "Market cap relative to free cash flow. Rewards companies that convert earnings to cash.", f"{WEIGHTS['price_fcf']:.0%}", "Lower = Better"],
        ["ROIC",             "Return on invested capital. Best single measure of business quality and competitive moat.", f"{WEIGHTS['roic']:.0%}", "Higher = Better"],
        ["GM Stability",     "Standard deviation of gross margin over 5 years. Stable margins signal pricing power.", f"{WEIGHTS['gm_stability']:.0%}", "Lower = Better"],
        ["Net Debt/EBITDA",  "Financial leverage. Negative means more cash than debt.", f"{WEIGHTS['net_debt_ebitda']:.0%}", "Lower = Better"],
    ]
    for row_idx, f in enumerate(factors, start=7):
        for col_idx, val in enumerate(f, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = make_fill(LIGHT_BLUE if row_idx % 2 == 0 else WHITE)
            cell.alignment = Alignment(horizontal="left" if col_idx <= 2 else "center", wrap_text=True)
            cell.border = make_border()
        ws.row_dimensions[row_idx].height = 30

    # Scoring explanation
    ws["A13"] = "Scoring"
    ws["A13"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
    ws["A14"] = "Each company is assigned a percentile rank (0–100) within the universe on each factor. The composite score is the weighted average of all five percentile ranks. Higher composite = more fundamentally attractive relative to peers."
    ws["A14"].font = Font(name="Arial", size=10)
    ws["A14"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A14:D14")
    ws.row_dimensions[14].height = 45

    # Color key
    ws["A16"] = "Score Key"
    ws["A16"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
    key = [("65–100", "Attractive", GREEN), ("40–64", "Neutral", YELLOW), ("0–39", "Unattractive", RED)]
    for i, (rng, label, color) in enumerate(key, start=17):
        ws.cell(row=i, column=1, value=rng).font = Font(name="Arial", size=10)
        cell = ws.cell(row=i, column=2, value=label)
        cell.font = Font(name="Arial", size=10)
        cell.fill = make_fill(color)
        cell.border = make_border()

    # Disclaimer
    ws["A21"] = "Disclaimer"
    ws["A21"].font = Font(name="Arial", bold=True, size=12, color=DARK_BLUE)
    ws["A22"] = "This model is a quantitative screening tool for educational and informational purposes only. Output does not constitute investment advice and should not be relied upon for investment decisions. Past factor performance does not guarantee future returns. Always conduct independent research before making any investment decision."
    ws["A22"].font = Font(name="Arial", size=10, italic=True, color="888888")
    ws["A22"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A22:D22")
    ws.row_dimensions[22].height = 60

def export(results):
    output_path = "output/screener_results.xlsx"
    os.makedirs("output", exist_ok=True)
    cols = ["ticker", "composite", "ev_ebit", "price_fcf",
            "roic", "gm_stability", "net_debt_ebitda", "n_factors"]

    # Round values
    results = results.copy()
    results["composite"] = results["composite"].round(1)
    for col in ["ev_ebit", "price_fcf", "roic", "gm_stability", "net_debt_ebitda"]:
        results[col] = results[col].round(2)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        results[cols].to_excel(writer, sheet_name="Ranked Output", index=False)
        results[cols].head(10).to_excel(writer, sheet_name="Top 10", index=False)

    # Post-process formatting
    wb = load_workbook(output_path)

    for sheet_name, title in [("Ranked Output", "All Companies — Ranked by Composite Score"),
                               ("Top 10", "Top 10 Most Attractive Companies")]:
        ws = wb[sheet_name]
        format_sheet(ws, results, title)

    add_methodology_tab(wb)

    # Set tab order
    wb.move_sheet("Methodology", offset=-wb.index(wb["Methodology"]))

    wb.save(output_path)
    print(f"Exported to {output_path}")

if __name__ == "__main__":
    print("Running screener...")
    df = build_factor_table(TICKERS)
    results = score(df)
    export(results)